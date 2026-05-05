from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def build_workbook_bytes(df: pd.DataFrame, title: str | None = None, notes: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja"

    row_cursor = 1
    if title:
        ws.cell(row=row_cursor, column=1, value=title).font = Font(bold=True, size=14)
        row_cursor += 2

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for col_idx, header in enumerate(df.columns, start=1):
        cell = ws.cell(row=row_cursor, column=col_idx, value=str(header))
        cell.font = Font(bold=True)
        cell.fill = header_fill
    row_cursor += 1

    for _, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_cursor, column=col_idx, value=value if pd.notna(value) else None)
        row_cursor += 1

    if notes:
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value=f"Notas: {notes}").font = Font(italic=True)

    for col_idx in range(1, max(len(df.columns), 1) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_workbook(data: dict, output_path: Path) -> Path:
    df = pd.DataFrame(data.get("rows") or [], columns=data.get("headers") or [])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(build_workbook_bytes(df, data.get("title"), data.get("notes")))
    return output_path
